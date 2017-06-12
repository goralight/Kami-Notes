import openpyxlimport openpyxl.worksheetfrom openpyxl import load_workbookfrom openpyxl.styles.borders import Sidefrom openpyxl.styles import PatternFill, Border, Alignmentimport datetimeimport osclass InitExcel:    def __init__(self, ConfigList):        """        Init the excel file. Basically create the file in the desired place with        the name generated via the user input. And capture the path for reuse.        Also populate the file with the required stuff for input is done. Things        like the core information about the file / headers        :param ConfigList: Get the options from the user needed to define core                           information        """        self.ConfigList = ConfigList        global CurrentWorkingExcelPath  # Need to change these from global        global BugLogNumber        BugLogNumber = ConfigList[5] + str(ConfigList[6])        Timestamp = datetime.datetime.now().strftime("%H%M-%Y.%m.%d")        self.CurrentWorkingExcelPath = ConfigList[4]+"/"+BugLogNumber+"-"+Timestamp+".xlsx"        # print self.ConfigList        BorderBottom = Border(bottom=Side(style="thin"))        BorderBottomRight = Border(bottom=Side(style="thin"), right=Side(style="thin"))        BorderRight = Border(right=Side(style="thin"))        PurpleFill = PatternFill(start_color="bdb0e8",  # No hash needed (#)                                 end_color="bdb0e8",                                 fill_type="solid")        Excel = openpyxl.Workbook()        WorkingSheet = Excel.get_sheet_by_name('Sheet')        WorkingSheet.title = BugLogNumber        WorkingSheet.column_dimensions["A"].width = 20.0  # 20=len of A3        WorkingSheet["A1"] = "Bug log for Jira entry:"        WorkingSheet["A2"] = "Reporters Name:"        WorkingSheet["A3"] = "Testing Environment:"        WorkingSheet["A4"] = "Charter Type:"        WorkingSheet["A5"] = "Time of Testing:"        WorkingSheet["A6"] = "Timer:"        WorkingSheet["A6"].border = BorderBottom        SetUpLength = float(len(ConfigList[8]))        WorkingSheet.column_dimensions["B"].width = SetUpLength        WorkingSheet["B1"] = BugLogNumber        WorkingSheet["B2"] = ConfigList[7]        WorkingSheet["B3"] = ConfigList[8]        WorkingSheet["B4"] = ConfigList[9]        WorkingSheet["B5"] = datetime.datetime.now().strftime("%H:%M:%S")        if "selected" in ConfigList[0]:            WorkingSheet["B6"] = datetime.timedelta(seconds=ConfigList[1]*60)        else:            WorkingSheet["B6"] = "Timer Disabled!"        WorkingSheet['B6'].alignment = Alignment(horizontal="left")        WorkingSheet["B6"].border = BorderBottomRight        WorkingSheet["B5"].border = BorderRight        WorkingSheet["B4"].border = BorderRight        WorkingSheet["B3"].border = BorderRight        WorkingSheet["B2"].border = BorderRight        WorkingSheet["B1"].border = BorderRight        WorkingSheet["A1"].fill = PurpleFill        WorkingSheet["A2"].fill = PurpleFill        WorkingSheet["A3"].fill = PurpleFill        WorkingSheet["A4"].fill = PurpleFill        WorkingSheet["A5"].fill = PurpleFill        WorkingSheet["A6"].fill = PurpleFill        WorkingSheet["B1"].fill = PurpleFill        WorkingSheet["B2"].fill = PurpleFill        WorkingSheet["B3"].fill = PurpleFill        WorkingSheet["B4"].fill = PurpleFill        WorkingSheet["B5"].fill = PurpleFill        WorkingSheet["B6"].fill = PurpleFill        WorkingSheet["A8"] = "Note Type"        WorkingSheet["B8"] = "Time Left"        WorkingSheet["C8"] = "Input"        WorkingSheet["D8"] = "TimeStamp"        WorkingSheet["A8"].border = BorderBottom        WorkingSheet["B8"].border = BorderBottom        WorkingSheet["C8"].border = BorderBottom        WorkingSheet["D8"].border = BorderBottom        WorkingSheet.column_dimensions["D"].width = 10.0        if not os.path.exists(ConfigList[4]):            os.makedirs(ConfigList[4])        Excel.save(self.CurrentWorkingExcelPath)class InputExcel:    def __init__(self, EntryInput, Row, CurrentWorkingExcelPath, EntryListInput, ConfigList):        """        Handles the loading of the created excel file and firing the InputRow        function to input to the file what the user inputted. It does have to        load the file every single time. Maybe a different approach? a function        called only once to handle loading?        :param EntryInput: Input widget which the user uses        :param Row: The row of which the next / start of the data is being inputted                    to. The starting value is 9 and will increment each time this                    is called        :param CurrentWorkingExcelPath: Path to the excel file being worked on        :param EntryListInput: List of what has been inputted so far. Used for history        """        self.EntryInput = EntryInput        self.Row = Row        self.CurrentWorkingExcelPath = CurrentWorkingExcelPath        self.EntryInputList = EntryListInput        self.ConfigList = ConfigList        self.Excel = load_workbook(filename=self.CurrentWorkingExcelPath)        self.WorkingSheet = self.Excel.get_sheet_by_name(BugLogNumber)        self.InputRow()    def InputRow(self):        """        Inputs to the row / columns the relative data being inputted. It will        also resize the column size each time something is inputted so everything        is readable and doesnt require manual resetting. Maybe do wrap text?        :return: EntryInputList, contains the list of everything been inputted                 so far. Used for the history button. Starts off at [] and just                 gets bigger and bigger        """        self.WorkingSheet["A" + str(self.Row)] = self.EntryInput[0]        if "selected" in self.ConfigList[0]:            self.WorkingSheet["B" + str(self.Row)] = self.EntryInput[1]        else:            self.WorkingSheet["B" + str(self.Row)] = "Timer Disabled!"        self.WorkingSheet["C" + str(self.Row)] = self.EntryInput[2]        self.WorkingSheet["D" + str(self.Row)] = self.EntryInput[3]        self.Row += 1        self.EntryInputList += [self.EntryInput[2]]        # Gets all the rows and figures out the width for the note column        for col in self.WorkingSheet.columns:            max_length = 0            column = col[0].column  # Get the column name            for cell in col:                try:  # Necessary to avoid error on empty cells                    if len(str(cell.value)) > max_length:                        max_length = len(cell.value)                except:                    pass            adjusted_width = (max_length + 2) * 1.03            self.WorkingSheet.column_dimensions[column].width = adjusted_width        self.Excel.save(self.CurrentWorkingExcelPath)        return self.EntryInputList